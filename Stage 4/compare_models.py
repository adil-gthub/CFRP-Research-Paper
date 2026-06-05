"""
compare_models.py
-----------------
Build a single Excel workbook comparing all four pushover models:
    M1 : Original RCC          (M30, ductile detailing)
    M2 : Degraded RCC          (M20, As × 0.85, 15-yr corrosion)
    M3 : CFRP retrofit – GF    (GF storey CFRP only)
    M4 : CFRP retrofit – Full  (every column wrapped + every beam laminated)

Reads the .npz files produced by each model's pushover script:
    outputs/disp_shear_rcc.npz          (M1)
    outputs/disp_shear_Degraded_rcc.npz (M2)
    outputs/disp_shear_cfrp.npz         (M3)
    outputs/disp_shear_cfrp_full.npz    (M4)

Run AFTER all four models have been executed:
    python compare_models.py

Output:
    outputs/Comparison_All_Models.xlsx
        Sheet 1 "Summary"       – peak +V, peak -V, initial K, drift cap, etc.
        Sheet 2 "Hysteresis"    – full (disp, shear) histories side-by-side
        Sheet 3 "Envelopes"     – peak (disp, shear) for each model
        Sheet 4 "Stiffness"     – secant K vs |disp| for each model
        Sheet 5 "CapacityCurve" – +ve envelope only (load vs top disp, mm)
"""

import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs')

# ============================================================================
# 1. LOAD ALL FOUR MODELS
# ============================================================================
MODELS = [
    ('M1_Original_RCC',     'disp_shear_rcc.npz'),
    ('M2_Degraded_RCC',     'disp_shear_Degraded_rcc.npz'),
    ('M3_CFRP_GF',          'disp_shear_cfrp.npz'),
    ('M4_CFRP_Full',        'disp_shear_cfrp_full.npz'),
]

data = {}
roof_height_mm = 17500.0   # head-room slab elevation = 17.5 m

print("Loading model results...")
for name, fname in MODELS:
    path = os.path.join(OUT_DIR, fname)
    if not os.path.isfile(path):
        print(f"  ⚠ {name}: file missing ({fname}) — skipping")
        continue
    d = np.load(path)
    disp  = d['disp']
    shear = d['shear']
    pidx  = d['peak_indices']
    env_disp  = disp[pidx]
    env_shear = shear[pidx]
    # Sort envelope by signed displacement for clean ordering
    order = np.argsort(env_disp)
    env_disp  = env_disp[order]
    env_shear = env_shear[order]
    # Stiffness vs |disp| (secant), positive disp side only
    pos      = env_disp > 0
    cap_disp = env_disp[pos]
    cap_V    = env_shear[pos]
    K_secant = np.where(cap_disp > 0, np.abs(cap_V) / cap_disp, 0.0)
    data[name] = dict(disp=disp, shear=shear,
                      env_disp=env_disp, env_shear=env_shear,
                      cap_disp=cap_disp, cap_V=cap_V, K=K_secant)
    print(f"  ✓ {name}: {len(disp)} points,  {len(pidx)} peaks,  "
          f"max +V = {np.max(env_shear):6.1f} kN")

if not data:
    raise RuntimeError("No model results found — run pushover scripts first.")

# ============================================================================
# 2. BUILD WORKBOOK
# ============================================================================
wb = Workbook()
ws = wb.active

HDR_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HDR_FILL   = PatternFill('solid', start_color='305496')
BLUE_INPUT = Font(name='Arial', color='0000FF')
BLACK_FORM = Font(name='Arial', color='000000')
GREEN_LINK = Font(name='Arial', color='008000')
BORDER     = Border(left=Side(style='thin', color='BFBFBF'),
                    right=Side(style='thin', color='BFBFBF'),
                    top=Side(style='thin', color='BFBFBF'),
                    bottom=Side(style='thin', color='BFBFBF'))
CENTER     = Alignment(horizontal='center', vertical='center')


def style_header_row(sheet, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = sheet.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

# ============================================================================
# SHEET 1 — SUMMARY  (single-row-per-model metrics)
# ============================================================================
ws.title = 'Summary'
ws['A1'] = '4-Model Cyclic Pushover Comparison'
ws['A1'].font = Font(name='Arial', size=14, bold=True)
ws.merge_cells('A1:I1')
ws['A2'] = ('M1 = Original (M30) · M2 = Degraded (M20+15%Ascorr) · '
            'M3 = CFRP GF-only · M4 = CFRP Full building')
ws.merge_cells('A2:I2')

headers = ['Model', 'Peak +V (kN)', 'Peak -V (kN)',
           'Disp @ peak +V (mm)', 'Disp @ peak -V (mm)',
           'Top-storey drift @ +V peak (%)',
           'Initial secant K (kN/mm)',
           'Ultimate disp (mm)', 'Notes']
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

NOTE_MAP = {
    'M1_Original_RCC': 'M30 conc, full As, Mander confined core',
    'M2_Degraded_RCC': 'M20 conc, As × 0.85, Mander confined core',
    'M3_CFRP_GF':      'M20 base + CFRP wrap GF cols + GF beam laminate',
    'M4_CFRP_Full':    'M20 base + CFRP wrap ALL cols + ALL beams laminated',
}

row = 5
for name in [n for n, _ in MODELS if n in data]:
    d = data[name]
    env_d  = d['env_disp']
    env_V  = d['env_shear']
    cap_d  = d['cap_disp']
    cap_V  = d['cap_V']
    K_init = d['K'][0] if len(d['K']) > 0 else 0.0  # secant at smallest +disp
    ws.cell(row=row, column=1, value=name).font = BLACK_FORM
    ws.cell(row=row, column=2, value=float(np.max(env_V))).font = BLUE_INPUT
    ws.cell(row=row, column=3, value=float(np.min(env_V))).font = BLUE_INPUT
    idx_max = int(np.argmax(env_V))
    idx_min = int(np.argmin(env_V))
    ws.cell(row=row, column=4, value=float(env_d[idx_max])).font = BLUE_INPUT
    ws.cell(row=row, column=5, value=float(env_d[idx_min])).font = BLUE_INPUT
    # drift % at +V peak via formula referencing col D (disp) and a fixed roof height cell
    ws.cell(row=row, column=6,
            value=f"=D{row}/$L$4").font = BLACK_FORM
    ws.cell(row=row, column=7, value=float(K_init)).font = BLUE_INPUT
    ws.cell(row=row, column=8, value=float(np.max(np.abs(env_d)))).font = BLUE_INPUT
    ws.cell(row=row, column=9, value=NOTE_MAP.get(name, '')).font = BLACK_FORM
    row += 1

# Roof height constant cell used by formula above
ws['K4'] = 'Roof height (mm):'
ws['L4'] = roof_height_mm
ws['L4'].font = BLUE_INPUT

# Number formats
for r in range(5, row):
    for c in (2, 3, 4, 5, 7, 8):
        ws.cell(row=r, column=c).number_format = '#,##0.00'
    ws.cell(row=r, column=6).number_format = '0.00%'

# Column widths
widths = [18, 14, 14, 18, 18, 22, 20, 16, 60]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Relative-comparison block
ws['A' + str(row + 2)] = 'Relative comparison (vs M1)'
ws['A' + str(row + 2)].font = Font(name='Arial', size=12, bold=True)
ws.cell(row=row+3, column=1, value='Model')
ws.cell(row=row+3, column=2, value='Peak +V / M1 +V')
ws.cell(row=row+3, column=3, value='Initial K / M1 K')
style_header_row(ws, row+3, 3)
# M1 row index (assume first)
m1_row = 5
out_r = row + 4
for r in range(5, row):
    ws.cell(row=out_r, column=1, value=f"=A{r}").font = GREEN_LINK
    ws.cell(row=out_r, column=2, value=f"=B{r}/B{m1_row}").font = BLACK_FORM
    ws.cell(row=out_r, column=2).number_format = '0.0%'
    ws.cell(row=out_r, column=3, value=f"=G{r}/G{m1_row}").font = BLACK_FORM
    ws.cell(row=out_r, column=3).number_format = '0.0%'
    out_r += 1

# ============================================================================
# SHEET 2 — HYSTERESIS (full disp + shear for each model side-by-side)
# ============================================================================
ws2 = wb.create_sheet('Hysteresis')
col = 1
for name in [n for n, _ in MODELS if n in data]:
    d = data[name]
    ws2.cell(row=1, column=col,   value=f"{name}: Disp (mm)")
    ws2.cell(row=1, column=col+1, value=f"{name}: Shear (kN)")
    for i, (u, V) in enumerate(zip(d['disp'], d['shear'])):
        ws2.cell(row=i+2, column=col,   value=float(u))
        ws2.cell(row=i+2, column=col+1, value=float(V))
    col += 2
style_header_row(ws2, 1, col - 1)
max_rows = max(len(data[n]['disp']) for n in [n for n, _ in MODELS if n in data])
for i in range(1, col):
    ws2.column_dimensions[get_column_letter(i)].width = 18
    for r in range(2, 2 + max_rows):
        cell = ws2.cell(row=r, column=i)
        if cell.value is not None:
            cell.number_format = '0.000'

# ============================================================================
# SHEET 3 — ENVELOPES  (peaks only)
# ============================================================================
ws3 = wb.create_sheet('Envelopes')
col = 1
for name in [n for n, _ in MODELS if n in data]:
    d = data[name]
    ws3.cell(row=1, column=col,   value=f"{name}: Env-Disp (mm)")
    ws3.cell(row=1, column=col+1, value=f"{name}: Env-Shear (kN)")
    for i, (u, V) in enumerate(zip(d['env_disp'], d['env_shear'])):
        ws3.cell(row=i+2, column=col,   value=float(u))
        ws3.cell(row=i+2, column=col+1, value=float(V))
    col += 2
style_header_row(ws3, 1, col - 1)
max_env_rows = max(len(data[n]['env_disp']) for n in [n for n, _ in MODELS if n in data])
for i in range(1, col):
    ws3.column_dimensions[get_column_letter(i)].width = 20
    for r in range(2, 2 + max_env_rows):
        cell = ws3.cell(row=r, column=i)
        if cell.value is not None:
            cell.number_format = '0.000'

# Add a comparison scatter chart of envelopes
chart3 = ScatterChart()
chart3.title    = 'Capacity Envelopes — All Models'
chart3.x_axis.title = 'Lateral Displacement (mm)'
chart3.y_axis.title = 'Lateral Load (kN)'
chart3.height = 12
chart3.width  = 22
c = 1
for name in [n for n, _ in MODELS if n in data]:
    n_pts = len(data[name]['env_disp'])
    x = Reference(ws3, min_col=c,   min_row=2, max_row=1 + n_pts)
    y = Reference(ws3, min_col=c+1, min_row=2, max_row=1 + n_pts)
    s = Series(y, x, title=name)
    s.smooth = False
    chart3.series.append(s)
    c += 2
ws3.add_chart(chart3, 'L2')

# ============================================================================
# SHEET 4 — STIFFNESS DEGRADATION  (secant K vs |+disp|)
# ============================================================================
ws4 = wb.create_sheet('Stiffness')
col = 1
for name in [n for n, _ in MODELS if n in data]:
    d = data[name]
    ws4.cell(row=1, column=col,   value=f"{name}: |Disp| (mm)")
    ws4.cell(row=1, column=col+1, value=f"{name}: Secant K (kN/mm)")
    for i, (u, K) in enumerate(zip(d['cap_disp'], d['K'])):
        ws4.cell(row=i+2, column=col,   value=float(u))
        ws4.cell(row=i+2, column=col+1, value=float(K))
    col += 2
style_header_row(ws4, 1, col - 1)
max_stiff_rows = max(len(data[n]['cap_disp']) for n in [n for n, _ in MODELS if n in data])
for i in range(1, col):
    ws4.column_dimensions[get_column_letter(i)].width = 22
    for r in range(2, 2 + max_stiff_rows):
        cell = ws4.cell(row=r, column=i)
        if cell.value is not None:
            cell.number_format = '0.000'

chart4 = ScatterChart()
chart4.title    = 'Stiffness Degradation — All Models'
chart4.x_axis.title = 'Top-Storey Displacement (mm)'
chart4.y_axis.title = 'Secant Stiffness (kN/mm)'
chart4.height = 12
chart4.width  = 22
c = 1
for name in [n for n, _ in MODELS if n in data]:
    n_pts = len(data[name]['cap_disp'])
    x = Reference(ws4, min_col=c,   min_row=2, max_row=1 + n_pts)
    y = Reference(ws4, min_col=c+1, min_row=2, max_row=1 + n_pts)
    s = Series(y, x, title=name)
    chart4.series.append(s)
    c += 2
ws4.add_chart(chart4, 'L2')

# ============================================================================
# SHEET 5 — CAPACITY CURVE  (load vs top-storey displacement, +ve side only)
# ============================================================================
ws5 = wb.create_sheet('CapacityCurve')
col = 1
for name in [n for n, _ in MODELS if n in data]:
    d = data[name]
    ws5.cell(row=1, column=col,   value=f"{name}: Top Disp (mm)")
    ws5.cell(row=1, column=col+1, value=f"{name}: Load (kN)")
    for i, (u, V) in enumerate(zip(d['cap_disp'], d['cap_V'])):
        ws5.cell(row=i+2, column=col,   value=float(u))
        ws5.cell(row=i+2, column=col+1, value=float(V))
    col += 2
style_header_row(ws5, 1, col - 1)
for i in range(1, col):
    ws5.column_dimensions[get_column_letter(i)].width = 22

chart5 = ScatterChart()
chart5.title    = 'Capacity Curve — Load vs Top-Storey Displacement'
chart5.x_axis.title = 'Top-Storey Displacement (mm)'
chart5.y_axis.title = 'Lateral Load (kN)'
chart5.height = 14
chart5.width  = 24
c = 1
for name in [n for n, _ in MODELS if n in data]:
    n_pts = len(data[name]['cap_disp'])
    x = Reference(ws5, min_col=c,   min_row=2, max_row=1 + n_pts)
    y = Reference(ws5, min_col=c+1, min_row=2, max_row=1 + n_pts)
    s = Series(y, x, title=name)
    chart5.series.append(s)
    c += 2
ws5.add_chart(chart5, 'L2')

# ============================================================================
# SAVE
# ============================================================================
out_path = os.path.join(OUT_DIR, 'Comparison_All_Models.xlsx')
wb.save(out_path)
print(f"\n✅ Comparison workbook written to {out_path}")
print("   Sheets: Summary | Hysteresis | Envelopes | Stiffness | CapacityCurve")
