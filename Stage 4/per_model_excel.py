"""
per_model_excel.py
------------------
Generate one Excel workbook per model from its .npz file.

Each workbook contains:
    Summary        : key metrics (peak +V, -V, K₀, drift, ductility)
    Hysteresis     : full disp/shear history + scatter chart
    Envelope       : peaks only + chart
    Stiffness      : secant K vs |disp| + chart
    CapacityCurve  : +ve envelope, load vs top-storey displacement + chart

Run AFTER pushover scripts have produced their npz files:
    python per_model_excel.py

Inputs (in ./outputs/):
    disp_shear_rcc.npz           → M1_Original_RCC.xlsx
    disp_shear_Degraded_rcc.npz  → M2_Degraded_RCC.xlsx
    disp_shear_cfrp.npz          → M3_CFRP_GF.xlsx
    disp_shear_cfrp_full.npz     → M4_CFRP_Full.xlsx
"""

import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'outputs')

MODELS = [
    # (display name, npz file, output xlsx, description)
    ('M1_Original_RCC',  'disp_shear_rcc.npz',
        'M1 — Original RCC (M30, full As, ductile detailing)'),
    ('M2_Degraded_RCC',  'disp_shear_Degraded_rcc.npz',
        'M2 — Degraded RCC (M20 after 15 yr, As × 0.85 corroded)'),
    ('M3_CFRP_GF',       'disp_shear_cfrp.npz',
        'M3 — CFRP retrofit, ground-floor columns + beams only'),
    ('M4_CFRP_Full',     'disp_shear_cfrp_full.npz',
        'M4 — CFRP retrofit, every storey wrapped & laminated'),
]

ROOF_HEIGHT_MM = 17500.0   # head-room slab elevation = 17.5 m

# ─── styles ─────────────────────────────────────────────────────────────────
ARIAL    = 'Arial'
HDR_FONT = Font(name=ARIAL, bold=True, color='FFFFFF', size=11)
HDR_FILL = PatternFill('solid', start_color='305496')
BLUE_IN  = Font(name=ARIAL, color='0000FF')                 # hardcoded inputs
BLACK_F  = Font(name=ARIAL, color='000000')                 # formulas
GREEN_L  = Font(name=ARIAL, color='008000')                 # cross-sheet links
CENTER   = Alignment(horizontal='center', vertical='center')
BORDER   = Border(left=Side(style='thin', color='BFBFBF'),
                  right=Side(style='thin', color='BFBFBF'),
                  top=Side(style='thin', color='BFBFBF'),
                  bottom=Side(style='thin', color='BFBFBF'))


def style_header(sheet, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = sheet.cell(row=row, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = (
            HDR_FONT, HDR_FILL, CENTER, BORDER)


def add_scatter(ws, title, x_label, y_label, x_col, y_col, n_rows,
                anchor='F2', smooth=True):
    chart = ScatterChart()
    chart.title         = title
    chart.x_axis.title  = x_label
    chart.y_axis.title  = y_label
    chart.height        = 12
    chart.width         = 20
    chart.legend        = None
    x = Reference(ws, min_col=x_col, min_row=2, max_row=1 + n_rows)
    y = Reference(ws, min_col=y_col, min_row=2, max_row=1 + n_rows)
    s = Series(y, x, title=title)
    s.smooth = smooth
    chart.series.append(s)
    ws.add_chart(chart, anchor)


# ─── per-model workbook builder ─────────────────────────────────────────────
def build_workbook(model_key, npz_path, description, out_xlsx):
    d     = np.load(npz_path)
    disp  = d['disp']
    shear = d['shear']
    pidx  = d['peak_indices']
    amps  = d['amps']
    ncyc  = int(d['n_cycles'])

    env_disp  = disp[pidx]
    env_shear = shear[pidx]
    # secant K at every peak (+ve disp only, sorted)
    pos       = env_disp > 0
    cap_disp  = env_disp[pos]
    cap_V     = env_shear[pos]
    K_secant  = np.where(cap_disp > 0, np.abs(cap_V) / cap_disp, 0.0)
    # stiffness sheet uses |disp| sort for clean curve
    stiff_d   = np.abs(env_disp)
    stiff_K   = np.where(stiff_d > 0, np.abs(env_shear) / stiff_d, 0.0)
    order     = np.argsort(stiff_d)
    stiff_d, stiff_K = stiff_d[order], stiff_K[order]

    wb = Workbook()

    # ── Summary ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = description
    ws['A1'].font = Font(name=ARIAL, size=14, bold=True)
    ws.merge_cells('A1:D1')

    rows = [
        ('Model', model_key, None),
        ('Protocol amplitudes (mm)', ', '.join(str(int(a)) for a in amps), None),
        ('Cycles per amplitude', ncyc, None),
        ('Total turning points', len(pidx), None),
        ('Total integration steps', len(disp), None),
        ('Roof height (mm)', ROOF_HEIGHT_MM, None),
        ('', '', None),
        ('Peak +V (kN)', float(env_shear.max()), 'hardcode'),
        ('Peak −V (kN)', float(env_shear.min()), 'hardcode'),
        ('Disp @ +V peak (mm)', float(env_disp[int(np.argmax(env_shear))]), 'hardcode'),
        ('Disp @ −V peak (mm)', float(env_disp[int(np.argmin(env_shear))]), 'hardcode'),
        ('Roof drift @ +V peak (%)', None, '=B11/B6*100'),
        ('Ultimate disp (mm)', float(np.max(np.abs(env_disp))), 'hardcode'),
        ('Initial secant K (kN/mm)', float(K_secant[0]) if len(K_secant) else 0.0, 'hardcode'),
        ('Final secant K (kN/mm)', float(K_secant[-1]) if len(K_secant) else 0.0, 'hardcode'),
        ('Stiffness retained (%)', None, '=B15/B14*100'),
    ]
    r = 3
    for label, value, kind in rows:
        ws.cell(row=r, column=1, value=label).font = BLACK_F
        if kind is None:
            ws.cell(row=r, column=2, value=value).font = BLUE_IN
        elif kind == 'hardcode':
            ws.cell(row=r, column=2, value=value).font = BLUE_IN
            ws.cell(row=r, column=2).number_format = '#,##0.00'
        else:  # formula
            ws.cell(row=r, column=2, value=kind).font = BLACK_F
            ws.cell(row=r, column=2).number_format = '0.00'
        r += 1
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # ── Hysteresis ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Hysteresis')
    ws2['A1'] = 'Disp (mm)'
    ws2['B1'] = 'Shear (kN)'
    style_header(ws2, 1, 2)
    for i, (u, V) in enumerate(zip(disp, shear)):
        ws2.cell(row=i+2, column=1, value=float(u)).number_format = '0.000'
        ws2.cell(row=i+2, column=2, value=float(V)).number_format = '0.000'
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 14
    add_scatter(ws2,
                f'Hysteretic Response — {model_key}',
                'Lateral Displacement (mm)', 'Lateral Load (kN)',
                x_col=1, y_col=2, n_rows=len(disp),
                anchor='D2', smooth=False)

    # ── Envelope ───────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Envelope')
    ws3['A1'] = 'Peak Disp (mm)'
    ws3['B1'] = 'Peak Shear (kN)'
    style_header(ws3, 1, 2)
    # sort by signed disp for clean envelope line
    order_env = np.argsort(env_disp)
    for i, (u, V) in enumerate(zip(env_disp[order_env], env_shear[order_env])):
        ws3.cell(row=i+2, column=1, value=float(u)).number_format = '0.000'
        ws3.cell(row=i+2, column=2, value=float(V)).number_format = '0.000'
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 18
    add_scatter(ws3,
                f'Capacity Envelope — {model_key}',
                'Lateral Displacement (mm)', 'Lateral Load (kN)',
                x_col=1, y_col=2, n_rows=len(env_disp),
                anchor='D2', smooth=False)

    # ── Stiffness ──────────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Stiffness')
    ws4['A1'] = '|Disp| (mm)'
    ws4['B1'] = 'Secant K (kN/mm)'
    style_header(ws4, 1, 2)
    for i, (u, K) in enumerate(zip(stiff_d, stiff_K)):
        ws4.cell(row=i+2, column=1, value=float(u)).number_format = '0.000'
        ws4.cell(row=i+2, column=2, value=float(K)).number_format = '0.000'
    ws4.column_dimensions['A'].width = 16
    ws4.column_dimensions['B'].width = 18
    add_scatter(ws4,
                f'Stiffness Degradation — {model_key}',
                'Lateral Displacement (mm)', 'Secant Stiffness (kN/mm)',
                x_col=1, y_col=2, n_rows=len(stiff_d),
                anchor='D2', smooth=False)

    # ── Capacity Curve  (load vs top-storey displacement, +ve side) ───────
    ws5 = wb.create_sheet('CapacityCurve')
    ws5['A1'] = 'Top-Storey Disp (mm)'
    ws5['B1'] = 'Lateral Load (kN)'
    style_header(ws5, 1, 2)
    sort_cap = np.argsort(cap_disp)
    for i, (u, V) in enumerate(zip(cap_disp[sort_cap], cap_V[sort_cap])):
        ws5.cell(row=i+2, column=1, value=float(u)).number_format = '0.000'
        ws5.cell(row=i+2, column=2, value=float(V)).number_format = '0.000'
    ws5.column_dimensions['A'].width = 22
    ws5.column_dimensions['B'].width = 18
    add_scatter(ws5,
                f'Capacity Curve — {model_key}',
                'Top-Storey Displacement (mm)', 'Lateral Load (kN)',
                x_col=1, y_col=2, n_rows=len(cap_disp),
                anchor='D2', smooth=False)

    wb.save(out_xlsx)


# ─── driver ────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("Generating per-model Excel workbooks...")
    for key, npz, desc in MODELS:
        path = os.path.join(OUT_DIR, npz)
        if not os.path.isfile(path):
            print(f"  ⚠ {key}: {npz} not found — skipping")
            continue
        out_xlsx = os.path.join(OUT_DIR, f'{key}.xlsx')
        build_workbook(key, path, desc, out_xlsx)
        d = np.load(path)
        print(f"  ✓ {key}: {len(d['disp'])} pts, "
              f"max +V = {d['shear'].max():7.1f} kN  →  {os.path.basename(out_xlsx)}")
    print("\n✅ Per-model workbooks complete.")
